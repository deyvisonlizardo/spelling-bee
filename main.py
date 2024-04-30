from tkinter import messagebox
import customtkinter
import random
import openpyxl
import pygame
from openpyxl.styles import PatternFill
import threading
import time
from PIL import Image
from gtts import gTTS
from playsound import playsound
import os


customtkinter.set_appearance_mode("System") # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("blue") # Themes: "blue" (standard), "green", "dark-blue"

class App(customtkinter.CTk):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        
        # configure window
        self.title("Spelling Bee 2024/1")
        self.geometry(f"{1280}x{720}")

        # configure grid layout
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(( 1, 2), weight=1)
        
        # sidebar frame - esquerda
        self.leftsidebar_frame = customtkinter.CTkFrame(self, width=140, corner_radius=0)
        self.leftsidebar_frame.grid(row=0, column=0, rowspan=5, sticky="nsew")
        self.leftsidebar_frame.grid_rowconfigure(4, weight=1)

        self.spelling_bee_logo = customtkinter.CTkImage(light_image=Image.open(os.path.join("images", "spellingbee-logo-black.png")),
                                  dark_image=Image.open(os.path.join("images", "spellingbee-logo-white.png")),
                                  size=(120, 66.6))
        self.spelling_bee_label = customtkinter.CTkLabel(self.leftsidebar_frame, image=self.spelling_bee_logo, text="")
        self.spelling_bee_label.grid(row=0, column=0, padx=20, pady=(20, 10))
        self.leftsidebar_button_1 = customtkinter.CTkButton(self.leftsidebar_frame, text="Reset NAMES", command=self.resetNames)
        self.leftsidebar_button_1.grid(row=1, column=0, padx=20, pady=10)
        self.leftsidebar_button_2 = customtkinter.CTkButton(self.leftsidebar_frame, text="Reset WORDS", command=self.resetWords)
        self.leftsidebar_button_2.grid(row=2, column=0, padx=20, pady=10)
        self.leftsidebar_button_3 = customtkinter.CTkButton(self.leftsidebar_frame, text="Open Second Window", command=self.openNewWindow)
        self.leftsidebar_button_3.grid(row=3, column=0, padx=20, pady=10)
        self.confirmation_response = customtkinter.CTkLabel(self.leftsidebar_frame, text="")
        self.confirmation_response.grid(row=4, column=0, padx=20, pady=(20, 10))

        self.appearance_mode_label = customtkinter.CTkLabel(self.leftsidebar_frame, text="Appearance Mode:", anchor="w")
        self.appearance_mode_label.grid(row=6, column=0, padx=20, pady=(10, 0))
        self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(self.leftsidebar_frame, values=["Light", "Dark"], command=self.change_appearance_mode_event)
        self.appearance_mode_optionemenu.grid(row=7, column=0, padx=20, pady=(10, 10))

        # main frame - middle
        self.middle_frame = customtkinter.CTkFrame(self)
        self.middle_frame.grid(row=0, column=1, rowspan=16, sticky="nsew")
        self.middle_frame.grid_rowconfigure((0, 3, 6, 9, 11, 14), weight=1)
        self.middle_frame.grid_columnconfigure((0, 4), weight=1)
        self.middle_frame.grid_columnconfigure((1, 2, 3), weight=0)
        
        self.middle_frame_word_title = customtkinter.CTkLabel(self.middle_frame, text="Word:", text_color="#38B6FF", font=customtkinter.CTkFont(size=20))
        self.middle_frame_word_title.grid(row=1, column=0, columnspan=5)
        self.middle_frame_word = customtkinter.CTkLabel(self.middle_frame, text="", font=customtkinter.CTkFont(family="New Order", size=70, weight="bold"))
        self.middle_frame_word.grid(row=2, column=0, columnspan=5)

        self.middle_frame_partofspeech= customtkinter.CTkLabel(self.middle_frame, text="Part of speech:", text_color="#38B6FF", font=customtkinter.CTkFont(size=20))
        self.middle_frame_partofspeech.grid(row=4, column=0, columnspan=5)
        self.middle_frame_speech = customtkinter.CTkLabel(self.middle_frame, text="", font=customtkinter.CTkFont(size=40, weight="bold"))
        self.middle_frame_speech.grid(row=5, column=0, columnspan=5)

        self.middle_frame_phrase_title = customtkinter.CTkLabel(self.middle_frame, text="Example in a sentence:", text_color="#38B6FF", font=customtkinter.CTkFont(size=20))
        self.middle_frame_phrase_title.grid(row=7, column=0, columnspan=5)
        self.middle_frame_wordphrase = customtkinter.CTkLabel(self.middle_frame, text="", font=customtkinter.CTkFont(size=40, weight="bold"))
        self.middle_frame_wordphrase.grid(row=8, column=0, columnspan=5)

        self.middle_colorbar = customtkinter.CTkLabel(self.middle_frame, text="------------------------------------------------------------------------------------------------------------------------------------------------------------------")
        self.middle_colorbar.grid(row=10, column=0, columnspan=5)

        self.middle_frame_student_title = customtkinter.CTkLabel(self.middle_frame, text="Student:", text_color="#38B6FF", font=customtkinter.CTkFont(size=20))
        self.middle_frame_student_title.grid(row=12, column=0, columnspan=5)
        self.middle_frame_student_name = customtkinter.CTkLabel(self.middle_frame, text="", font=customtkinter.CTkFont(size=40, weight="bold"))
        self.middle_frame_student_name.grid(row=13, column=0, columnspan=5)
        

        # main frame - middle buttons
        self.middle_frame_button_1 = customtkinter.CTkButton(self.middle_frame, text="Approved", fg_color="green", font=customtkinter.CTkFont(weight="bold"), command=self.approved)
        self.middle_frame_button_1.grid(row=15, column=1, padx=10, pady=30)
        self.middle_frame_button_2 = customtkinter.CTkButton(self.middle_frame, text="Out", fg_color="red", font=customtkinter.CTkFont(weight="bold"), command=self.out)
        self.middle_frame_button_2.grid(row=15, column=2, padx=10, pady=30)
        self.middle_frame_button_3 = customtkinter.CTkButton(self.middle_frame, text="Absent", text_color="#000000", fg_color="yellow", font=customtkinter.CTkFont(weight="bold"), command=self.absent)
        self.middle_frame_button_3.grid(row=15, column=3, padx=10, pady=30)
        self.middle_frame_button_4 = customtkinter.CTkButton(self.middle_frame, text="Next Student", fg_color="#3A4751", font=customtkinter.CTkFont(weight="bold"), command=self.nextStudent)
        self.middle_frame_button_4.grid(row=15, column=4, padx=10, pady=30)

        # sidebar fram - direita
        self.rightsidebar_frame = customtkinter.CTkFrame(self, width=150, corner_radius=0)
        self.rightsidebar_frame.grid(row=0, column=2, rowspan=10, sticky="nsew")
        self.rightsidebar_frame.grid_rowconfigure(6, weight=1)
        
        self.timer = customtkinter.CTkLabel(self.rightsidebar_frame, width=150, height=40, text="Timer", font=customtkinter.CTkFont(size=30))
        self.timer.grid(row=0, column=0, padx=20, pady=(30, 20))

        self.timer_entry = customtkinter.CTkEntry(self.rightsidebar_frame, placeholder_text="Enter time in seconds")
        self.timer_entry.grid(row=1, column=0, padx=20, pady=(20, 10))

        self.timer_confirm = customtkinter.CTkButton(self.rightsidebar_frame, text="Start Timer", command=self.start_timer)
        self.timer_confirm.grid(row=2, column=0, padx=20, pady=(0, 10))

        self.timer_stop = customtkinter.CTkButton(self.rightsidebar_frame, text="Stop Timer", command=self.stop_timer)
        self.timer_stop.grid(row=3, column=0, padx=20, pady=(0, 10))

        self.next_word = customtkinter.CTkButton(self.rightsidebar_frame, text="Word pronunciation", fg_color="#3A4751", command=self.getWord)
        self.next_word.grid(row=4, column=0, padx=10, pady=(30, 0))

        self.next_word = customtkinter.CTkButton(self.rightsidebar_frame, text="Next Word", fg_color="#3A4751", command=self.nextWord)
        self.next_word.grid(row=5, column=0, padx=10, pady=10)

        self.round = customtkinter.CTkLabel(self.rightsidebar_frame, width=150, height=40, text="", font=customtkinter.CTkFont(size=30))
        self.round.grid(row=7, column=0, padx=20, pady=(30, 20))

        self.next_round = customtkinter.CTkButton(self.rightsidebar_frame, text="Next Round", text_color="#000000", fg_color="orange", font=customtkinter.CTkFont(weight="bold"), command=self.nextRound)
        self.next_round.grid(row=8, column=0, padx=20, pady=(0, 10))

        self.icbeu_logo = customtkinter.CTkImage(light_image=Image.open(os.path.join("images", "icbeu-logo.png")),
                                  dark_image=Image.open(os.path.join("images", "icbeu-logo.png")),
                                  size=(120, 60.3))
        self.icbeulogo_label = customtkinter.CTkLabel(self.rightsidebar_frame, image=self.icbeu_logo, text="")
        self.icbeulogo_label.grid(row=9, column=0, padx=20, pady=(20, 50))


        # set default values
        self.appearance_mode_optionemenu.set("Dark")
        self.new_window = None
        self.round_count = 1
        self.round.configure(text="Round " + f"{self.round_count}")
        self.stop_thread = False
        self.current_thread = None
        

    # functions
    def change_appearance_mode_event(self, new_appearance_mode: str):
        customtkinter.set_appearance_mode(new_appearance_mode)
        if new_appearance_mode == "Light":

            text_color_light = "#000000"
            self.middle_frame_word_title.configure(text_color=f"{text_color_light}")
            self.middle_frame_partofspeech.configure(text_color=f"{text_color_light}")
            self.middle_frame_phrase_title.configure(text_color=f"{text_color_light}")
            self.middle_frame_student_title.configure(text_color=f"{text_color_light}")

        if new_appearance_mode == "Dark":

            text_color_dark="#38B6FF"
            self.middle_frame_word_title.configure(text_color=f"{text_color_dark}")
            self.middle_frame_partofspeech.configure(text_color=f"{text_color_dark}")
            self.middle_frame_phrase_title.configure(text_color=f"{text_color_dark}")
            self.middle_frame_student_title.configure(text_color=f"{text_color_dark}")

    def openNewWindow(self):
        if self.new_window is None or not self.new_window.winfo_exists():
            self.new_window = customtkinter.CTkToplevel(self)
            self.new_window.title("Spelling Bee 2024/1")
            self.new_window.geometry(f"{1280}x{720}")
            self.new_window.grid_columnconfigure(1, weight=1)
            self.new_window.grid_rowconfigure(1, weight=1)

            self.second_window_frame = customtkinter.CTkFrame(self.new_window)
            self.second_window_frame.grid(row=0, column=0, rowspan=3, columnspan=3, sticky="nsew")
            self.second_window_frame.grid_rowconfigure((2, 5), weight=1)
            self.second_window_frame.grid_rowconfigure(3, weight=0)
            self.second_window_frame.grid_columnconfigure(1, weight=1)
            self.second_window_frame.grid_columnconfigure((0, 2), weight=0)

            self.spelling_bee_logo_sw = customtkinter.CTkImage(light_image=Image.open(os.path.join("images", "spellingbee-logo-black.png")),
                                  dark_image=Image.open(os.path.join("images", "spellingbee-logo-white.png")),
                                  size=(120, 66.6))
            self.spelling_bee_label_sw = customtkinter.CTkLabel(self.second_window_frame, width=150, image=self.spelling_bee_logo_sw, text="")
            self.spelling_bee_label_sw.grid(row=0, column=0, padx=20, pady=(20, 10))


            self.second_window_student_title = customtkinter.CTkLabel(self.second_window_frame, text="Student:", text_color="#38B6FF", font=customtkinter.CTkFont(size=20))
            self.second_window_student_title.grid(row=0, column=0, columnspan=3, pady=(30, 10))
            self.second_window_student_name = customtkinter.CTkLabel(self.second_window_frame, text="", font=customtkinter.CTkFont(size=40, weight="bold"))
            self.second_window_student_name.grid(row=1, column=0, columnspan=3)
            student = self.middle_frame_student_name.cget("text")
            self.second_window_student_name.configure(text=f"{student}")

            self.second_window_word_title = customtkinter.CTkLabel(self.second_window_frame, text="Word:", text_color="#38B6FF", font=customtkinter.CTkFont(size=20))
            self.second_window_word_title.grid(row=3, column=0, columnspan=3, pady=(30, 10))

            self.new_window_label = customtkinter.CTkLabel(self.second_window_frame, text="", font=customtkinter.CTkFont(family="New Order", size=100, weight="bold"))
            self.new_window_label.grid(row=4, column=0, columnspan=3)
            word = self.middle_frame_word.cget("text")
            self.new_window_label.configure(text=f"{word}")

            self.second_window_timer = customtkinter.CTkLabel(self.second_window_frame, width=150, height=40, text="Timer", font=customtkinter.CTkFont(size=30))
            self.second_window_timer.grid(row=0, column=2, rowspan=2, padx=20, pady=(30, 20))

            self.icbeu_logo_sw = customtkinter.CTkImage(light_image=Image.open(os.path.join("images", "icbeu-logo.png")),
                                  dark_image=Image.open(os.path.join("images", "icbeu-logo.png")),
                                  size=(120, 60.3))
            self.icbeulogo_label_sw = customtkinter.CTkLabel(self.second_window_frame, image=self.icbeu_logo_sw, text="")
            self.icbeulogo_label_sw.grid(row=6, column=2, padx=20, pady=(20, 50))
            
        else:
            self.new_window.focus()


    def resetNames(self):
        try:
            # Carregar o arquivo Excel
            nome_arquivo = os.path.join("database", "students.xlsx")
            workbook = openpyxl.load_workbook(nome_arquivo)
            sheet = workbook.active

            # Percorrer todas as linhas da coluna B
            no_fill = PatternFill(fill_type=None)

            # Iterate over all rows in column B (column B corresponds to column number 2)
            for row in sheet.iter_rows(min_col=2, max_col=2):
                for cell in row:
                    cell.fill = no_fill
                    # Salvar as mudanças de volta no arquivo
            
            workbook.save(nome_arquivo)

            #Mensagem de confirmação na tela
            self.confirmation_response.configure(text="Success!\nNames reseted", font=customtkinter.CTkFont(size=18, weight="bold"), text_color="#7CFC00")
            workbook.close()

        except Exception as e:
            self.confirmation_response.configure(text="Failed\nto reset", font=customtkinter.CTkFont(size=18, weight="bold"), text_color="red")


    def resetWords(self):
        try:
            # Carregar o arquivo Excel
            nome_arquivo = os.path.join("database", "words.xlsx")
            workbook = openpyxl.load_workbook(nome_arquivo)
            sheet = workbook.active

            # Percorrer todas as linhas da coluna E
            for row in sheet.iter_rows(min_row=1, min_col=5, max_col=5):
                cell = row[0]
                cell.value = None

            # Salvar as mudanças de volta no arquivo
            workbook.save(nome_arquivo)
            
            #Mensagem de confirmação na tela
            self.confirmation_response.configure(text="Success!\nWords reseted", font=customtkinter.CTkFont(size=18, weight="bold"), text_color="#7CFC00")
            workbook.close()

        except Exception as e:
            self.confirmation_response.configure(text="Failed\nto reset", font=customtkinter.CTkFont(size=18, weight="bold"), text_color="red")


    def getWord(self):
        word = self.middle_frame_word.cget("text") #Get the word showed in the interface

        try:
            t1 = threading.Thread(target=self.wordPronunciation, args=[word])
            t1.start()
        
        except ValueError:
            print(ValueError)

            return

    def wordPronunciation(self, word):
        # Create a gTTS object
        tts = gTTS(text=word, lang='en')
    
        # Use os.path.join to ensure the correct path format across different OS
        file_path = os.path.join("sounds", f"{word}.mp3")

        # Save the speech as a temporary file
        tts.save(file_path)

        # Play the speech
        playsound(file_path)

        time.sleep(0.5)  # Wait for a second to ensure the file is not being used by the player
        
        try:
            os.remove(file_path)
        except PermissionError:
            print(f"Failed to delete {file_path}, file might still be in use.")



    def nextWord(self):
        try:
            workbook = openpyxl.load_workbook(os.path.join("database", "words.xlsx"))
            sheet = workbook.active
            
            available_rows = []

            count = 1
            # Check every row in column A (starts from 1 as per openpyxl's convention)
            for row in sheet.iter_rows(min_col=1, max_col=1):
                for cell in row:
                    if cell.value:  # Check if the cell has content
                        count += 1
            
            for row in range(1, count):
                if sheet.cell(row=row, column=5).value != "Sorteada":
                    available_rows.append(row)

            if not available_rows:
                messagebox.showinfo("Aviso", "Todas as palavras já foram sorteadas.")
                workbook.close()
                return
            
            random_number = random.choice(available_rows)
            
            cell_value_e = sheet.cell(row=random_number, column=5).value
            
            if cell_value_e != "Sorteada":
                cell_value_b = sheet.cell(row=random_number, column=2).value
                cell_value_c = sheet.cell(row=random_number, column=3).value
                cell_value_d = sheet.cell(row=random_number, column=4).value

                self.update_display(cell_value_b, cell_value_c, cell_value_d)
                sheet.cell(row=random_number, column=5, value="Sorteada")
                workbook.save(os.path.join("database", "words.xlsx"))

            
            workbook.close()

            self.getWord()
            self.start_timer()

        except Exception as e:
            print(f"Erro ao acessar o arquivo: {e}")


    def nextStudent(self):
        try:
            workbook = openpyxl.load_workbook(os.path.join("database", "students.xlsx"))
            sheet = workbook.active
            
            available_rows = []

            count = 1
            # Check every row in column A (starts from 1 as per openpyxl's convention)
            for row in sheet.iter_rows(min_col=1, max_col=1):
                for cell in row:
                    if cell.value:  # Check if the cell has content
                        count += 1

            for row in range(1, count):
                if sheet.cell(row=row, column=2).fill.fill_type is None:
                    available_rows.append(row)

            if not available_rows:
                messagebox.showinfo("Aviso", "Todas os alunos já foram sorteadas.")
                workbook.close()
                return
            
            random_number = random.choice(available_rows)

            cell_color = sheet.cell(row=random_number, column=2).fill.fill_type

            if cell_color is None:
                cell_name = sheet.cell(row=random_number, column=1).value
                
                self.update_student(cell_name)
                self.stop_timer()
            
            workbook.close()
            self.confirmation_response.configure(text="")
        
        except Exception as e:
            print(f"Erro ao acessar o arquivo: {e}")


    def approved(self):
        try:
            # Load workbook and select the active worksheet
            workbook = openpyxl.load_workbook(os.path.join("database", "students.xlsx"))
            sheet = workbook.active
            name_to_search = self.middle_frame_student_name.cget("text")

            # Define a green fill pattern
            color_fill = PatternFill(start_color="7CFC00", end_color="7CFC00", fill_type="solid")
            name_found = False

            # Check every cell in column A (column A corresponds to column number 1)
            for row in sheet.iter_rows(min_col=1, max_col=1):
                for cell in row:
                    # If the name is found in column A
                    if cell.value == name_to_search:
                        # Color the corresponding cell in column B (of the same row) with green
                        sheet.cell(row=cell.row, column=2).fill = color_fill
                        name_found = True
            
            if name_found:
                self.confirmation_response.configure(text="Student\nis Approved", font=customtkinter.CTkFont(size=20, weight="bold"), text_color="#7CFC00")
            else:
                self.confirmation_response.configure(text="Name not found", font=customtkinter.CTkFont(size=20, weight="bold"), text_color="red")
                print(color_fill.fgColor.rgb)
 
            # Save the workbook
            workbook.save(os.path.join("database", "students.xlsx"))
            workbook.close()
            
        except Exception as e:
            self.confirmation_response.configure(text="Failed\nto open database", font=customtkinter.CTkFont(size=20, weight="bold"), text_color="red")


    def out(self):
        try:
            # Load workbook and select the active worksheet
            workbook = openpyxl.load_workbook(os.path.join("database", "students.xlsx"))
            sheet = workbook.active
            name_to_search = self.middle_frame_student_name.cget("text")

            # Define a red fill pattern
            color_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            name_found = False

            # Check every cell in column A (column A corresponds to column number 1)
            for row in sheet.iter_rows(min_col=1, max_col=1):
                for cell in row:
                    # If the name is found in column A
                    if cell.value == name_to_search:
                        # Color the corresponding cell in column B (of the same row) with red
                        sheet.cell(row=cell.row, column=2).fill = color_fill
                        name_found = True
            
            if name_found:
                self.confirmation_response.configure(text="Student is Out", font=customtkinter.CTkFont(size=20, weight="bold"), text_color="#FF0000")
            else:
                self.confirmation_response.configure(text="Name not found", font=customtkinter.CTkFont(size=20, weight="bold"), text_color="red")
                print(color_fill.fgColor.rgb)

            # Save the workbook
            workbook.save(os.path.join("database", "students.xlsx"))
            workbook.close()
            
        except Exception as e:
            self.confirmation_response.configure(text="Failed\nto open database", font=customtkinter.CTkFont(size=20, weight="bold"), text_color="red")


    def absent(self):
        try:
            # Load workbook and select the active worksheet
            workbook = openpyxl.load_workbook(os.path.join("database", "students.xlsx"))
            sheet = workbook.active
            name_to_search = self.middle_frame_student_name.cget("text")

            # Define a red fill pattern
            color_fill = PatternFill(start_color="FFEA00", end_color="FFEA00", fill_type="solid")
            name_found = False

            # Check every cell in column A (column A corresponds to column number 1)
            for row in sheet.iter_rows(min_col=1, max_col=1):
                for cell in row:
                    # If the name is found in column A
                    if cell.value == name_to_search:
                        # Color the corresponding cell in column B (of the same row) with red
                        sheet.cell(row=cell.row, column=2).fill = color_fill
                        name_found = True
            
            if name_found:
                self.confirmation_response.configure(text="Student\nis Absent", font=customtkinter.CTkFont(size=20, weight="bold"), text_color="#FFEA00")
            else:
                self.confirmation_response.configure(text="Name not found", font=customtkinter.CTkFont(size=20, weight="bold"), text_color="red")
                print(color_fill.fgColor.rgb)

            # Save the workbook
            workbook.save(os.path.join("database", "students.xlsx"))
            workbook.close()
            
        except Exception as e:
            self.confirmation_response.configure(text="Failed\nto open database", font=customtkinter.CTkFont(size=20, weight="bold"), text_color="red")


    def nextRound(self):
        try:
            # Carregar o arquivo Excel
            nome_arquivo = os.path.join("database", "students.xlsx")
            workbook = openpyxl.load_workbook(nome_arquivo)
            sheet = workbook.active
            winner = 0

            # Percorrer todas as linhas da coluna B
            for row in sheet.iter_rows(min_col=2, max_col=2):
                for cell in row:
                    # If the cell's background color is green (Red RGB: 00FF0000)
                    if cell.fill.start_color.rgb == "007CFC00":
                        winner = winner + 1
                        winners_name = sheet.cell(row=cell.row, column=1).value
                        cell.fill = openpyxl.styles.PatternFill(fill_type=None)  # Set to no fill
        
            if winner == 1:
                self.middle_frame_word_title.configure(text="")
                self.middle_frame_word.configure(text="")

                self.middle_frame_partofspeech.configure(text="WINNER")
                self.middle_frame_speech.configure(text=f"{winners_name}")

                self.middle_frame_phrase_title.configure(text="")
                self.middle_frame_wordphrase.configure(text="")

                self.middle_frame_student_title.configure(text="")
                self.middle_frame_student_name.configure(text="")

                self.middle_frame_student_name.configure(text="")

                if self.new_window:
                    self.second_window_student_title.configure(text="WINNER")
                    self.second_window_student_name.configure(text="")

                    self.new_window_label.configure(text=f"{winners_name}")

                messagebox.showinfo("WINNER!!",f" {winners_name}")#, f"WE HAVE A WINNER:\n{winner_student}")

                workbook.close()
                
                return

            workbook.save(nome_arquivo)
            workbook.close()


            nome_arquivo = os.path.join("database", "words.xlsx")
            workbook = openpyxl.load_workbook(nome_arquivo)
            sheet = workbook.active

            # Percorrer todas as linhas da coluna E
            for row in sheet.iter_rows(min_row=1, min_col=5, max_col=5):
                cell = row[0]
                cell.value = None

            # Salvar as mudanças de volta no arquivo
            workbook.save(nome_arquivo)
            workbook.close()

            #Remove todas as informações da tela (name, word, part of..., etc)
            self.middle_frame_student_name.configure(text="")
            self.middle_frame_word.configure(text="")
            self.middle_frame_speech.configure(text="")
            self.middle_frame_wordphrase.configure(text="")

            if self.new_window:
                    self.second_window_student_title.configure(text="")
                    self.second_window_student_name.configure(text="")
                    self.second_window_word_title.configure(text="")
                    self.new_window_label.configure(text="")

            #Mensagem de confirmação na tela
            self.confirmation_response.configure(text="Success!\nNext round!", font=customtkinter.CTkFont(size=20, weight="bold"), text_color="#7CFC00")
            self.round_count += 1
            self.round.configure(text="Round " + f"{self.round_count}")
            

        except Exception as e:
            self.confirmation_response.configure(text="Close\nthe database!", font=customtkinter.CTkFont(size=20, weight="bold"), text_color="red")


    def update_display(self, word, speech, phrase):
        self.middle_frame_word.configure(text=f"{word}")
        self.middle_frame_speech.configure(text=f"{speech}")
        self.middle_frame_wordphrase.configure(text=f"{phrase}")

        if self.new_window:
            self.new_window_label.configure(text=f"{word}")


    def update_student(self, student_name):
        self.middle_frame_student_name.configure(text=f"{student_name}")
        
        self.middle_frame_word.configure(text="")
        self.middle_frame_speech.configure(text="")
        self.middle_frame_wordphrase.configure(text="")

        if self.new_window:            
            self.second_window_student_name.configure(text=f"{student_name}")

            self.new_window_label.configure(text="")


    def start_timer(self):

        if self.current_thread:
            self.stop_thread = True
            self.current_thread.join()

        try:
            seconds = int(self.timer_entry.get())
            self.stop_thread = False
            self.current_thread = threading.Thread(target=self.countdown, args=(seconds,))
            self.current_thread.start()

        except ValueError:
            self.timer_var.set("Invalid input")
        
        
    def stop_timer(self):

        if self.current_thread:
            self.stop_thread = True
            self.current_thread.join()
            self.timer.configure(text="Timer\nstopped!")

            if self.new_window:
                self.second_window_timer.configure(text="Timer\nstopped")


    def countdown(self, seconds):

        while seconds >= 0 and not self.stop_thread:
           
            mins, sec = divmod(seconds, 60)
            timer_format = '{:02d}:{:02d}'.format(mins, sec)
            self.timer.configure(text=timer_format, font=customtkinter.CTkFont(size=30, weight="bold"))
            
            if self.new_window:
                self.second_window_timer.configure(text=timer_format, font=customtkinter.CTkFont(size=30, weight="bold"))

            time.sleep(1)
            seconds -= 1


        if not self.stop_thread:
            self.timer.configure(text="Time's up!", font=customtkinter.CTkFont(size=30, weight="bold"))
            
            if self.new_window:
                self.second_window_timer.configure(text="Time's up!", font=customtkinter.CTkFont(size=30, weight="bold"))

            pygame.mixer.init(44100, -16,2,2048)
            pygame.mixer.music.load(os.path.join("sounds", "alarm.mp3"))
            pygame.mixer.music.play()
    

if __name__ == "__main__":
    app = App()
    app.mainloop()